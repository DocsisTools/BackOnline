import os
import sqlite3
from tkinter import *
from interface_stage import scm_before_input, scm_before_data, scm_after_data, query_output_box, casa_status_label, cisco_status_label
from disable_xls import *
from casa_cmts_class import CasaScm
from cisco_functions import *
import xlwt
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl import Workbook, load_workbook
from tkinter import messagebox
from tkinter import filedialog
import logging


logging.basicConfig(level=logging.INFO, filename = 'log.log',filemode="w",
                    format= "%(asctime)s - %(levelname)s - %(message)s")

def connect_db():
    global db_check
    current_dir = os.path.dirname(__file__)
    path_check = os.path.join(current_dir, 'scm_db_dev.db')
    db_check=os.path.isfile(path_check)
    
    global conn,c
    
    conn = sqlite3.connect(path_check)
    c = conn.cursor()
    
    if db_check == False:
        logging.info('DATABASE NOT DETECTED, CREATEING DATABASE...')
        c.execute('''CREATE TABLE cisco_scm_core_before_full(
            mac_before text,
            ip_before text,
            interface_before text,
            state_before text,
            sid text,
            rxpwr text,
            timingoff text,
            numcpe text,
            dip text
            )''')

        c.execute('''CREATE TABLE cisco_scm_core_after_full(
            mac_after text,
            ip_after text,
            interface_after text,
            state_after text,
            sid text,
            rxpwr text,
            timingoff text,
            numcpe text,
            dip text
            )''')

        c.execute('''CREATE TABLE casa_scm_core_before_full(
            mac_before text,
            ip_before text,
            us_interface_before text,
            ds_interface_before text,
            state_before text,
            sid text,
            rxpwr text,
            timingoffset text,
            numcpe text,
            bpi text
            )''')

        c.execute('''CREATE TABLE casa_scm_core_after_full(
            mac_after text,
            ip_after text,
            us_interface_after text,
            ds_interface_after text,
            state_after text,
            sid text,
            rxpwr text,
            timingoffset text,
            numcpe text,
            bpi text
            )''')
        pass

    else: 
        logging.info('CONNECTION TO DATABASE GOOD')
        return db_check 

def casa_scm_before():
    casa_status_label.configure(text = 'LOADING...')
    connect_db()
    clear_record_after='DELETE FROM casa_scm_core_before_full'
    c.execute(clear_record_after)

    datainput=scm_before_input.get(1.0,END).splitlines()
    for lines in datainput:
        try:
            cm_create=CasaScm.from_string_input(lines)
            
            c.execute("INSERT INTO casa_scm_core_before_full VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                      (cm_create.mac, cm_create.ip, cm_create.us_interface, cm_create.ds_interface,cm_create.status,cm_create.sid,cm_create.rxpwr,cm_create.timingoff,cm_create.numcpe,cm_create.bpi))
            

        except Exception as e:

            logging.debug(e)
            pass
        
    conn.commit()


    c.execute("SELECT DISTINCT * FROM casa_scm_core_before_full")
    rows_total = c.fetchall()
    total_count = len(rows_total)
    
    c.execute('''
        SELECT DISTINCT mac_before, state_before
        FROM casa_scm_core_before_full
        WHERE state_before = 'online(pt)'
    ''')
    
    rows_wb_onl = c.fetchall()
    online_count = len(rows_wb_onl)
    
    c.execute('''
        SELECT DISTINCT mac_before, state_before,us_interface_before
        FROM casa_scm_core_before_full
        WHERE us_interface_before like '%/%/0*' or us_interface_before like '%/%/0#' and state_before != "offline"
    ''')
    
    rows_wb_onl = c.fetchall()
    bonding_count = len(rows_wb_onl)
    
    c.execute('''
        SELECT DISTINCT mac_before, state_before,us_interface_before
        FROM casa_scm_core_before_full
        WHERE state_before like "offline"
    ''')
    
    rows_off = c.fetchall()
    offline_count = len(rows_off)
    
    c.execute('''
        SELECT DISTINCT mac_before, state_before,us_interface_before
        FROM casa_scm_core_before_full
        WHERE state_before like "init%"
    ''')
    
    rows_init = c.fetchall()
    init_count = len(rows_init)
    
    c.execute('''
        SELECT DISTINCT mac_before, state_before,us_interface_before
        FROM casa_scm_core_before_full
        WHERE state_before !="online(pt)"
    ''')
        
    rows_not_onl = c.fetchall()
    not_online_count = len(rows_not_onl)
    
    c.execute('''
        SELECT DISTINCT mac_before, state_before,us_interface_before
        FROM casa_scm_core_before_full
        WHERE state_before ="online(pt)" AND numcpe='0'
    ''')
    
    rows_onl_0cpe = c.fetchall()
    onl_0cpe_count = len(rows_onl_0cpe)
    
    c.execute('''
        SELECT DISTINCT mac_before, state_before,us_interface_before
        FROM casa_scm_core_before_full
        WHERE us_interface_before like '%/%/0*' or us_interface_before like '%/%/0#'
    ''')
    
    rows_usbond_onl = c.fetchall()
    usbonding_count = len(rows_usbond_onl)
    
    c.execute('''
        SELECT DISTINCT mac_before, state_before,ds_interface_before
        FROM casa_scm_core_before_full
        WHERE ds_interface_before like '%/%/%*' or ds_interface_before like '%/%/%#'
    ''')
    
    rows_dsbond_onl = c.fetchall()
    dsbonding_count = len(rows_dsbond_onl)

    scm_before_data.delete(1.0,END)
    scm_before_data.insert(END,'TOTAL_BEFORE: \n')
    scm_before_data.insert(END,total_count)
    scm_before_data.insert(END,'\nREG_BEFORE: \n')
    scm_before_data.insert(END,online_count)
    scm_before_data.insert(END,'\nBONDING_BEFORE: \n')
    scm_before_data.insert(END,bonding_count)
    scm_before_data.insert(END,'\nOFFLINE_BEFORE: \n')
    scm_before_data.insert(END,offline_count)
    scm_before_data.insert(END,'\nINIT_BEFORE: \n')
    scm_before_data.insert(END,init_count)
    scm_before_data.insert(END,'\nNOT_ONL_BEFORE: \n')
    scm_before_data.insert(END,not_online_count)
    scm_before_data.insert(END,'\nONL_0CPE_BEFORE: \n')
    scm_before_data.insert(END,onl_0cpe_count)
    scm_before_data.insert(END,'\nUS_BOND_BEFORE: \n')
    scm_before_data.insert(END,usbonding_count)
    scm_before_data.insert(END,'\nDS_BOND_BEFORE: \n')
    scm_before_data.insert(END,dsbonding_count)

    logging.info('TOTAL_BEFORE: ' + str(total_count))
    logging.info('REG_BEFORE: ' + str(online_count))
    logging.info('BONDING_BEFORE: ' + str(bonding_count))
    logging.info('OFFLINE_BEFORE: ' + str(offline_count))
    logging.info('INIT_BEFORE: ' + str(init_count))
    logging.info('NOT_ONL_BEFORE: ' + str(not_online_count))
    logging.info('ONL_0CPE_BEFORE: ' + str(onl_0cpe_count))
    logging.info('US_BOND_BEFORE: ' + str(usbonding_count))
    logging.info('DS_BOND_BEFORE: ' + str(dsbonding_count))


    casa_status_label.configure(text = 'CASA_SCM_BEFORE_LOADED',fg_color='#EF820D')
    cisco_status_label.configure(text = '', fg_color='transparent')
    disable_xls_buttons()
    
def casa_scm_after():
    casa_status_label.configure(text = 'LOADING...')
    connect_db()

    clear_record_after='DELETE FROM casa_scm_core_after_full'
    c.execute(clear_record_after)

    datainput=scm_before_input.get(1.0,END).splitlines()
    total_count=0
    
    for lines in datainput:
        try:
            cm_create=CasaScm.from_string_input(lines)
            
            c.execute("INSERT INTO casa_scm_core_after_full VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                      (cm_create.mac, cm_create.ip, cm_create.us_interface, cm_create.ds_interface,cm_create.status,cm_create.sid,cm_create.rxpwr,cm_create.timingoff,cm_create.numcpe,cm_create.bpi))
            
        except Exception as e:
            pass
    
    conn.commit()
    
    c.execute("SELECT DISTINCT * FROM casa_scm_core_after_full")
    rows_total = c.fetchall()
    total_count = len(rows_total)
    
    c.execute('''
        SELECT DISTINCT mac_after, state_after
        FROM casa_scm_core_after_full
        WHERE state_after = 'online(pt)'
    ''')
    
    rows_onl = c.fetchall()
    online_count = len(rows_onl)
    
    c.execute('''
        SELECT DISTINCT mac_after, state_after,us_interface_after
        FROM casa_scm_core_after_full
        WHERE us_interface_after like '%/%/0*' or us_interface_after like '%/%/0#' and state_after != "offline"
    ''')
    
    rows_wb_onl = c.fetchall()
    bonding_count = len(rows_wb_onl)
    
    c.execute('''
        SELECT DISTINCT mac_after, state_after,us_interface_after
        FROM casa_scm_core_after_full
        WHERE state_after like "offline"
    ''')
    
    rows_off = c.fetchall()
    offline_count = len(rows_off)
    
    c.execute('''
        SELECT DISTINCT mac_after, state_after,us_interface_after
        FROM casa_scm_core_after_full
        WHERE state_after like "init%"
    ''')
    
    rows_init = c.fetchall()
    init_count = len(rows_init)
    
    c.execute('''
        SELECT DISTINCT mac_after, state_after,us_interface_after
        FROM casa_scm_core_after_full
        WHERE state_after !="online(pt)"
    ''')
        
    rows_not_onl = c.fetchall()
    not_online_count = len(rows_not_onl)
    
    c.execute('''
        SELECT DISTINCT mac_after, state_after,us_interface_after
        FROM casa_scm_core_after_full
        WHERE state_after ="online(pt)" AND numcpe='0'
    ''')
    
    rows_onl_0cpe = c.fetchall()
    onl_0cpe_count = len(rows_onl_0cpe)
    
    c.execute('''
        SELECT DISTINCT mac_after, state_after,us_interface_after
        FROM casa_scm_core_after_full
        WHERE us_interface_after like '%/%/0*' or us_interface_after like '%/%/0#'
    ''')
    
    rows_usbond_onl = c.fetchall()
    usbonding_count = len(rows_usbond_onl)
    
    c.execute('''
        SELECT DISTINCT mac_after, state_after,ds_interface_after
        FROM casa_scm_core_after_full
        WHERE ds_interface_after like '%/%/%*' or ds_interface_after like '%/%/%#'
    ''')
    
    rows_dsbond_onl = c.fetchall()
    dsbonding_count = len(rows_dsbond_onl)

    scm_after_data.delete(1.0,END)
    scm_after_data.insert(END,'TOTAL_AFTER: \n')
    scm_after_data.insert(END,total_count)
    scm_after_data.insert(END,'\nREG_AFTER: \n')
    scm_after_data.insert(END,online_count)
    scm_after_data.insert(END,'\nBONDING_AFTER: \n')
    scm_after_data.insert(END,bonding_count)
    scm_after_data.insert(END,'\nOFFLINE_AFTER: \n')
    scm_after_data.insert(END,offline_count)
    scm_after_data.insert(END,'\nINIT_AFTER: \n')
    scm_after_data.insert(END,init_count)
    scm_after_data.insert(END,'\nNOT_ONL_AFTER: \n')
    scm_after_data.insert(END,not_online_count)
    scm_after_data.insert(END,'\nONL_0CPE_AFTER: \n')
    scm_after_data.insert(END,onl_0cpe_count)
    scm_after_data.insert(END,'\nUS_BOND_AFTER: \n')
    scm_after_data.insert(END,usbonding_count)
    scm_after_data.insert(END,'\nDS_BOND_AFTER: \n')
    scm_after_data.insert(END,dsbonding_count)

    
    logging.info('TOTAL_AFTER: ' + str(total_count))
    logging.info('REG_AFTER: ' + str(online_count))
    logging.info('BONDING_AFTER: ' + str(bonding_count))
    logging.info('OFFLINE_AFTER: ' + str(offline_count))
    logging.info('INIT_AFTER: ' + str(init_count))
    logging.info('NOT_ONL_AFTER: ' + str(not_online_count))
    logging.info('ONL_0CPE_AFTER: ' + str(onl_0cpe_count))
    logging.info('US_BOND_AFTER: ' + str(usbonding_count))
    logging.info('DS_BOND_AFTER: ' + str(dsbonding_count))

    casa_status_label.configure(text = 'CASA_SCM_AFTER_LOADED',fg_color='#EF820D')
    cisco_status_label.configure(text = '', fg_color='transparent')
    disable_xls_buttons()
    
def casa_scm_browse_pre():
    scm_before_input.delete(1.0, END)
    casa_scm_pre_input = filedialog.askopenfilename(initialdir = "/",
                                          title = "Select a File",
                                          filetypes = (("Text files",
                                                        "*.txt*"),
                                                       ("all files",
                                                        "*.*")))
    scm_file_input = open(casa_scm_pre_input, 'r')
    
    for lines in scm_file_input:
        printline=lines
        scm_before_input.insert(END, printline)
        
    scm_file_input.close()
    casa_scm_before()
    pass

def casa_scm_browse_posle():
    scm_before_input.delete(1.0, END)
    casa_scm_posle_input = filedialog.askopenfilename(initialdir = "/",
                                          title = "Select a File",
                                          filetypes = (("Text files",
                                                        "*.txt*"),
                                                       ("all files",
                                                        "*.*")))
    scm_file_input = open(casa_scm_posle_input, 'r')
    
    for lines in scm_file_input:
        printline=lines
        scm_before_input.insert(END, printline)
        
    scm_file_input.close()
    casa_scm_after()
    pass


##################################################################################################################


def casa_compare_new_modems_correction():
    connect_db()
    c.execute('''
    SELECT DISTINCT mac_after, state_before, state_after
    FROM casa_scm_core_after_full
    LEFT JOIN casa_scm_core_before_full
    ON casa_scm_core_after_full.mac_after = casa_scm_core_before_full.mac_before
    WHERE state_before IS NULL
    ORDER BY state_after,state_before
    ''')
    rows_compare = c.fetchall()
    output = ""
    for row in rows_compare:
        mac_before = row[0] if row[0] is not None else "None"
        state_before = row[1] if row[1] is not None else "None"
        state_after = row[2] if row[2] is not None else "None"
        
        output += f"{mac_before:<16} {state_before:<10} {state_after:<5}\n"

    query_output_box.insert(1.0, output)
    disable_xls_buttons()
    pass

def casa_compare_new_modems_correction_detail():
    connect_db()
    c.execute('''
    SELECT DISTINCT mac_after, state_before, state_after, us_interface_before, us_interface_after,ds_interface_before, ds_interface_after
    FROM casa_scm_core_after_full
    LEFT JOIN casa_scm_core_before_full
    ON casa_scm_core_after_full.mac_after = casa_scm_core_before_full.mac_before
    WHERE state_before IS NULL
    ORDER BY state_after,state_before
    ''')
    rows_compare = c.fetchall()
    output = ""
    for row in rows_compare:
        mac_after = row[0] if row[0] is not None else "None"
        state_before = row[1] if row[1] is not None else "None"
        state_after = row[2] if row[2] is not None else "None"
        us_interface_before = row[3] if row[3] is not None else "None"
        us_interface_after = row[4] if row[4] is not None else "None"
        ds_interface_before = row[5] if row[5] is not None else "None"
        ds_interface_after = row[6] if row[6] is not None else "None"
        output += f"{mac_after:<16} {state_before:<12} {state_after:<12}{us_interface_before:<12} {us_interface_after:<12}{ds_interface_before:<10} {ds_interface_after:<10}\n"

    scm_before_input.insert(1.0, output)
    disable_xls_buttons()
    pass



def casa_compare_different_state():
    connect_db()
    c.execute('''
    SELECT DISTINCT mac_before, state_before, state_after
    FROM casa_scm_core_before_full
    LEFT JOIN casa_scm_core_after_full
    ON casa_scm_core_before_full.mac_before = casa_scm_core_after_full.mac_after
    WHERE (state_before != state_after) OR state_after IS NULL
    ORDER BY state_after,state_before
    ''')
    rows_compare = c.fetchall()
    output = ""
    for row in rows_compare:
        mac_before = row[0] if row[0] is not None else "None"
        state_before = row[1] if row[1] is not None else "None"
        state_after = row[2] if row[2] is not None else "None"
        
        output += f"{mac_before:<16} {state_before:<10} {state_after:<5}\n"


    query_output_box.delete(1.0, END)
    query_output_box.insert(END, output)
    
    casa_compare_new_modems_correction()
    casa_status_label.configure(text = 'CASA_BEFORE!=AFTER_LOADED',fg_color='#E55D2F')
    cisco_status_label.configure(text = '', fg_color='transparent')
    disable_xls_buttons()
    

def casa_compare_onl_pre():
    connect_db()
    c.execute('''
    SELECT DISTINCT mac_before, state_before, state_after
    FROM casa_scm_core_before_full
    LEFT JOIN casa_scm_core_after_full
    ON casa_scm_core_before_full.mac_before = casa_scm_core_after_full.mac_after
    WHERE (state_before = 'online(pt)' AND (state_before != state_after or state_after IS NULL))
    ORDER BY state_after,state_before
    ''')
    rows_compare = c.fetchall()
    output = ""
    for row in rows_compare:
        mac_before = row[0] if row[0] is not None else "None"
        state_before = row[1] if row[1] is not None else "None"
        state_after = row[2] if row[2] is not None else "None"
        
        output += f"{mac_before:<16} {state_before:<10} {state_after:<5}\n"


    query_output_box.delete(1.0, END)
    query_output_box.insert(END, output)
    casa_status_label.configure(text = 'CASA_ONLINE_BEFORE_LOADED',fg_color='#E55D2F')
    cisco_status_label.configure(text = '', fg_color='transparent')
    disable_xls_buttons()
    pass


def casa_compare_bonding():
    connect_db()
    c.execute('''
    SELECT DISTINCT mac_before, state_before, state_after, us_interface_before, us_interface_after
    FROM casa_scm_core_before_full
    LEFT JOIN casa_scm_core_after_full
    ON casa_scm_core_before_full.mac_before = casa_scm_core_after_full.mac_after
    WHERE (us_interface_before like '%/%/0*' or us_interface_before like '%/%/0#') and (us_interface_after not like '%/%/0*')
    ORDER BY state_after
    ''')
    rows_compare = c.fetchall()
    output = ""
    for row in rows_compare:
        mac_before = row[0] if row[0] is not None else "None"
        state_before = row[1] if row[1] is not None else "None"
        state_after = row[2] if row[2] is not None else "None"
        us_interface_before = row[3] if row[3] is not None else "None"
        us_interface_after = row[4] if row[4] is not None else "None"
        
        output += f"{mac_before:<16}{state_after:<10} {us_interface_before:<10} {us_interface_after:<10}\n"


    query_output_box.delete(1.0, END)
    query_output_box.insert(END, output)
    casa_status_label.configure(text = 'CASA_BONDING_DIFF_LOADED',fg_color='#E55D2F')
    cisco_status_label.configure(text = '', fg_color='transparent')
    disable_xls_buttons()

def casa_compare_all():
    from operator import itemgetter
    connect_db()
    c.execute('''
    SELECT DISTINCT mac_before, state_before, state_after, us_interface_before, us_interface_after
    FROM casa_scm_core_before_full
    LEFT JOIN casa_scm_core_after_full
    ON casa_scm_core_before_full.mac_before = casa_scm_core_after_full.mac_after
    ORDER BY state_after,state_before
    ''')
    rows_compare = c.fetchall()
    output = ""
    for row in rows_compare:
        mac_before = row[0] if row[0] is not None else "None"
        state_before = row[1] if row[1] is not None else "None"
        state_after = row[2] if row[2] is not None else "None"
        us_interface_before = row[3] if row[3] is not None else "None"
        us_interface_after = row[4] if row[4] is not None else "None"
        
        output += f"{mac_before:<16} {state_before:<10} {state_after:<10}\n"

    query_output_box.delete(1.0, END)
    query_output_box.insert(END, output)
    casa_compare_new_modems_correction()
    casa_status_label.configure(text = 'CASA_COMPARE_ALL_LOADED',fg_color='#E55D2F')
    cisco_status_label.configure(text = '', fg_color='transparent')
    disable_xls_buttons()
    
def casa_compare_all_detail():
    from operator import itemgetter
    connect_db()
    c.execute('''
    SELECT DISTINCT mac_before, state_before, state_after, us_interface_before, us_interface_after,ds_interface_before, ds_interface_after
    FROM casa_scm_core_before_full
    LEFT JOIN casa_scm_core_after_full
    ON casa_scm_core_before_full.mac_before = casa_scm_core_after_full.mac_after
    ORDER BY state_after,state_before
    ''')
    rows_compare = c.fetchall()
    output = ""
    for row in rows_compare:
        mac_before = row[0] if row[0] is not None else "None"
        state_before = row[1] if row[1] is not None else "None"
        state_after = row[2] if row[2] is not None else "None"
        us_interface_before = row[3] if row[3] is not None else "None"
        us_interface_after = row[4] if row[4] is not None else "None"
        ds_interface_before = row[5] if row[5] is not None else "None"
        ds_interface_after = row[6] if row[6] is not None else "None"
        
        
        output += f"{mac_before:<16} {state_before:<12} {state_after:<12}{us_interface_before:<12} {us_interface_after:<12}{ds_interface_before:<10} {ds_interface_after:<10}\n"

    scm_before_input.delete(1.0, END)
    scm_before_input.insert(END, output)
    casa_compare_new_modems_correction_detail()
    casa_status_label.configure(text = 'CASA_DETAIL_LOADED',fg_color='#E55D2F')
    cisco_status_label.configure(text = '', fg_color='transparent')

    enable_casa_xls_button()

def export_xl_casa():
    
    current_folder = os.getcwd() 
    file_path = os.path.join(current_folder, "casa_analysis.xlsx")
    rows = scm_before_input.get(1.0, "end-1c").splitlines()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'CASA CMTS MODEM ANALYSIS'
    # Column names
    column_names = ["MAC", "STATUS_BEFORE", "STATUS_AFTER","US_BEFORE","US_AFTER","DS_BEFORE","DS_AFTER"]
    for col_index, col_name in enumerate(column_names, start=1):
        col_letter = get_column_letter(col_index)
        sheet[col_letter + '1'] = col_name
        sheet[col_letter + '1'].font = Font(bold=True)
    # Iterate over the rows and split each row into columns
    for row_index, row in enumerate(rows, start=2):
        columns = row.split()
        for col_index, col_value in enumerate(columns, start=1):
            col_letter = get_column_letter(col_index)
            sheet[col_letter + str(row_index)] = col_value
            
    sheet.column_dimensions['A'].width = 18
    sheet.column_dimensions['B'].width = 18        
    sheet.column_dimensions['C'].width = 18
    sheet.column_dimensions['D'].width = 18
    sheet.column_dimensions['E'].width = 18
    sheet.column_dimensions['F'].width = 18
    sheet.column_dimensions['G'].width = 18
    sheet.column_dimensions['I'].width = 18
    sheet.column_dimensions['J'].width = 12
    sheet.column_dimensions['L'].width = 18
    sheet.column_dimensions['M'].width = 12

    
    data_pre=scm_before_data.get(1.0, "end-1c").splitlines()
    data_pre_dict = {}
    for i in range(0, len(data_pre), 2):
        key = data_pre[i]
        value = data_pre[i + 1]
        data_pre_dict[key] = value

    data_posle=scm_after_data.get(1.0, "end-1c").splitlines()
    data_posle_dict = {}
    for i in range(0, len(data_posle), 2):
        key = data_posle[i]
        value = data_posle[i + 1]
        data_posle_dict[key] = value
    
    for row_index, (key, value) in enumerate(data_pre_dict.items(), start=2):
        sheet.cell(row=row_index, column=9, value=key)
        sheet.cell(row=row_index, column=10, value=value)
    
    for row_index, (key, value) in enumerate(data_posle_dict.items(), start=2):
        sheet.cell(row=row_index, column=12, value=key)
        sheet.cell(row=row_index, column=13, value=value)


    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    fill = PatternFill(fill_type='solid', fgColor='E5F993')
    for row in sheet.iter_rows(min_row=2, max_row=10, min_col=9, max_col=10):
        for cell in row:
            cell.fill = fill
            cell.border = border
            
    fill = PatternFill(fill_type='solid', fgColor='F9DC5C')
    for row in sheet.iter_rows(min_row=2, max_row=10, min_col=12, max_col=13):
        for cell in row:
            cell.fill = fill
            cell.border = border
    
    try:
        os.remove(file_path)
        workbook.save(file_path)
        messagebox.showinfo('EXPORT SUCCESS', "OUTPUT SAVED TO casa_analysis.xlsx")
        
    except Exception as e:
        workbook.save(file_path)
        messagebox.showinfo('EXPORT SUCCESS', "OUTPUT SAVED TO casa_analysis.xlsx")
        pass
    

    os.startfile(file_path)
    
    pass

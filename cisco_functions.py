import os
import sqlite3
from tkinter import *
#from interface import scm_before_input, scm_before_data, scm_after_data, query_output_box, status_label
from interface_stage import scm_before_input, scm_before_data, scm_after_data, query_output_box, cisco_status_label,casa_status_label
from disable_xls import *
from cisco_cbr_class import CiscoScm
from cisco_functions import *
import xlwt
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl import Workbook, load_workbook
from tkinter import messagebox
from tkinter import filedialog
import logging


logging.basicConfig(level=logging.INFO, filename = 'log.log',filemode="w",
                    format= "%(asctime)s - %(levelname)s - %(message)s")

def connect_db():
    global db_check
    current_dir = os.path.dirname(__file__)
    path_check = os.path.join(current_dir, 'scm_db_dev.db')
    db_check=os.path.isfile(path_check)
    
    global conn,c
    
    conn = sqlite3.connect(path_check)
    c = conn.cursor()
    
    if db_check == False:
        logging.info('DATABASE NOT DETECTED, CREATEING DATABASE...')
        c.execute('''CREATE TABLE cisco_scm_core_before_full(
            mac_before text,
            ip_before text,
            interface_before text,
            state_before text,
            sid text,
            rxpwr text,
            timingoff text,
            numcpe text,
            dip text
            )''')

        c.execute('''CREATE TABLE cisco_scm_core_after_full(
            mac_after text,
            ip_after text,
            interface_after text,
            state_after text,
            sid text,
            rxpwr text,
            timingoff text,
            numcpe text,
            dip text
            )''')

        c.execute('''CREATE TABLE casa_scm_core_before_full(
            mac_before text,
            ip_before text,
            us_interface_before text,
            ds_interface_before text,
            state_before text,
            sid text,
            rxpwr text,
            timingoffset text,
            numcpe text,
            bpi text
            )''')

        c.execute('''CREATE TABLE casa_scm_core_after_full(
            mac_after text,
            ip_after text,
            us_interface_after text,
            ds_interface_after text,
            state_after text,
            sid text,
            rxpwr text,
            timingoffset text,
            numcpe text,
            bpi text
            )''')
        pass

    else: 
        logging.info('CONNECTION TO DATABASE GOOD')
        return db_check 

def boban():
    connect_db()


def input_scm_before():
    cisco_status_label.configure(text = 'LOADING...',fg_color='#00A0DA')
    connect_db()
    if db_check:
        clear_record_before='DELETE FROM cisco_scm_core_before_full'
        c.execute(clear_record_before)

        datainput=scm_before_input.get(1.0,END).splitlines()
        
        for lines in datainput:
            try:
                cm_create=CiscoScm.from_string_input(lines)

                c.execute("INSERT INTO cisco_scm_core_before_full VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                          (cm_create.mac, cm_create.ip, cm_create.interface, cm_create.state, cm_create.sid, cm_create.rxpwr, cm_create.timingoff, cm_create.numcpe, cm_create.dip))

            except Exception as e:
                logging.debug(e)
                pass
            
        conn.commit()

        c.execute("SELECT DISTINCT * FROM cisco_scm_core_before_full")
        rows_total = c.fetchall()
        total_count = len(rows_total)

        c.execute('''
            SELECT DISTINCT mac_before, state_before
            FROM cisco_scm_core_before_full
            WHERE state_before = 'w-online(pt)' OR state_before = 'p-online(pt)'
        ''')

        rows_wb_onl = c.fetchall()
        wideband_count = len(rows_wb_onl)

        c.execute('''
            SELECT DISTINCT mac_before, state_before
            FROM cisco_scm_core_before_full
            WHERE state_before = 'online(pt)'
        ''')

        rows_stb_onl = c.fetchall()
        stb_count = len(rows_stb_onl)

        c.execute('''
            SELECT DISTINCT mac_before, state_before
            FROM cisco_scm_core_before_full
            WHERE state_before = 'offline'
        ''')

        rows_off = c.fetchall()
        off_count = len(rows_off)

        c.execute('''
            SELECT DISTINCT mac_before, state_before
            FROM cisco_scm_core_before_full
            WHERE state_before like 'init%'
        ''')

        rows_init = c.fetchall()
        init_count = len(rows_init)

        scm_before_data.delete(1.0,END)
        scm_before_data.insert(END,'TOTAL_BEFORE: \n')
        scm_before_data.insert(END,total_count)
        scm_before_data.insert(END,'\nWIDEBAND_BEFORE: \n')
        scm_before_data.insert(END,wideband_count)
        scm_before_data.insert(END,'\nSTB_BEFORE: \n')
        scm_before_data.insert(END,stb_count)
        scm_before_data.insert(END,'\nOFFLINE_BEFORE: \n')
        scm_before_data.insert(END,off_count)
        scm_before_data.insert(END,'\nINIT_BEFORE: \n')
        scm_before_data.insert(END,init_count)
        
        logging.info('TOTAL_BEFORE: ' + str(total_count))
        logging.info('WIDEBAND_BEFORE: ' + str(wideband_count))
        logging.info('STB_BEFORE: ' + str(stb_count))
        logging.info('OFFLINE_BEFORE: ' + str(off_count))
        logging.info('INIT_BEFORE: ' + str(init_count))
        
        conn.close()
        cisco_status_label.configure(text = 'CISCO_SCM_BEFORE_LOADED',fg_color='#00A0DA')
        casa_status_label.configure(text = '', fg_color='transparent')
        
        disable_xls_buttons()
        
    else:
        connect_db()
        input_scm_before()
    
def input_scm_after():
    cisco_status_label.configure(text = 'LOADING...',fg_color='#00A0DA')
    connect_db()
    if db_check:
        clear_record_after='DELETE FROM cisco_scm_core_after_full'
        c.execute(clear_record_after)

        datainput=scm_before_input.get(1.0,END).splitlines()
        for lines in datainput:
            try:
                cm_create=CiscoScm.from_string_input(lines)

                c.execute("INSERT INTO cisco_scm_core_after_full VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                          (cm_create.mac, cm_create.ip, cm_create.interface, cm_create.state, cm_create.sid, cm_create.rxpwr, cm_create.timingoff, cm_create.numcpe, cm_create.dip))

            except Exception as e:
                pass

        conn.commit()

        c.execute("SELECT DISTINCT * FROM cisco_scm_core_after_full")
        rows_total = c.fetchall()
        total_count = len(rows_total)

        c.execute('''
            SELECT DISTINCT mac_after, state_after
            FROM cisco_scm_core_after_full
            WHERE state_after = 'w-online(pt)' OR state_after = 'p-online(pt)'
        ''')

        rows_wb_onl = c.fetchall()
        wideband_count = len(rows_wb_onl)

        c.execute('''
            SELECT DISTINCT mac_after, state_after
            FROM cisco_scm_core_after_full
            WHERE state_after = 'online(pt)'
        ''')

        rows_stb_onl = c.fetchall()
        stb_count = len(rows_stb_onl)

        c.execute('''
            SELECT DISTINCT mac_after, state_after
            FROM cisco_scm_core_after_full
            WHERE state_after = 'offline'
        ''')

        rows_off = c.fetchall()
        off_count = len(rows_off)

        c.execute('''
            SELECT DISTINCT mac_after, state_after
            FROM cisco_scm_core_after_full
            WHERE state_after like 'init%'
        ''')

        rows_init = c.fetchall()
        init_count = len(rows_init)

        scm_after_data.delete(1.0,END)
        scm_after_data.insert(END,'TOTAL_AFTER: \n')
        scm_after_data.insert(END,total_count)
        scm_after_data.insert(END,'\nWIDEBAND_AFTER: \n')
        scm_after_data.insert(END,wideband_count)
        scm_after_data.insert(END,'\nSTB_AFTER: \n')
        scm_after_data.insert(END,stb_count)
        scm_after_data.insert(END,'\nOFFLINE_AFTER: \n')
        scm_after_data.insert(END,off_count)
        scm_after_data.insert(END,'\nINIT_AFTER: \n')
        scm_after_data.insert(END,init_count)
        
        logging.info('TOTAL_AFTER: ' + str(total_count))
        logging.info('WIDEBAND_AFTER: ' + str(wideband_count))
        logging.info('STB_AFTER: ' + str(stb_count))
        logging.info('OFFLINE_AFTER: ' + str(off_count))
        logging.info('INIT_AFTER: ' + str(init_count))
        
        conn.close()
        cisco_status_label.configure(text = 'CISCO_SCM_AFTER_LOADED',fg_color='#00A0DA')
        casa_status_label.configure(text = '', fg_color='transparent')
        disable_xls_buttons()
    else:
        connect_db()
        input_scm_after()
  
def cisco_scm_browse_before():
    scm_before_input.delete(1.0, END)
    cisco_scm_pre_input = filedialog.askopenfilename(initialdir = "/",
                                          title = "Select a File",
                                          filetypes = (("Text files",
                                                        "*.txt*"),
                                                       ("all files",
                                                        "*.*")))
    scm_file_input = open(cisco_scm_pre_input, 'r')
    
    for lines in scm_file_input:
        printline=lines
        scm_before_input.insert(END, printline)
        
    scm_file_input.close()
    input_scm_before()
    pass

def cisco_scm_browse_after():

    scm_before_input.delete(1.0, END)
    cisco_scm_posle_input = filedialog.askopenfilename(initialdir = "/",
                                          title = "Select a File",
                                          filetypes = (("Text files",
                                                        "*.txt*"),
                                                       ("all files",
                                                        "*.*")))
    scm_file_input = open(cisco_scm_posle_input, 'r')
    
    for lines in scm_file_input:
        printline=lines
        scm_before_input.insert(END, printline)
        
    scm_file_input.close()
    input_scm_after()
    pass
     

########################################################################################
########################################################################################

def cisco_compare_new_modems_correction():
    connect_db()
    c.execute('''
    SELECT DISTINCT mac_after, state_before, state_after
    FROM cisco_scm_core_after_full
    LEFT JOIN cisco_scm_core_before_full
    ON cisco_scm_core_after_full.mac_after = cisco_scm_core_before_full.mac_before
    WHERE state_before IS NULL
    ORDER BY state_after,state_before
    ''')
    rows_compare = c.fetchall()
    output = ""
    for row in rows_compare:
        mac_before = row[0] if row[0] is not None else "None"
        state_before = row[1] if row[1] is not None else "None"
        state_after = row[2] if row[2] is not None else "None"
        
        output += f"{mac_before:<16} {state_before:<10} {state_after:<5}\n"

    query_output_box.insert(1.0, output)
    
    pass

def cisco_join_compare():
    connect_db()
    c.execute('''
    SELECT DISTINCT mac_before, state_before, state_after
    FROM cisco_scm_core_before_full
    LEFT JOIN cisco_scm_core_after_full
    ON cisco_scm_core_before_full.mac_before = cisco_scm_core_after_full.mac_after
    ORDER BY state_before,state_after
    ''')
    
    rows_compare = c.fetchall()
    
    output = ""
    for row in rows_compare:
        mac_before = row[0] if row[0] is not None else "None"
        state_before = row[1] if row[1] is not None else "None"
        state_after = row[2] if row[2] is not None else "None"
        
        output += f"{mac_before:<16} {state_before:<12} {state_after:<12}\n"

    query_output_box.delete(1.0, END)
    query_output_box.insert(END, output)
    cisco_status_label.configure(text = 'CISCO_COMPARE_ALL_LOADED',fg_color='#054FA8')
    casa_status_label.configure(text = '', fg_color='transparent')
    cisco_compare_new_modems_correction()
    disable_xls_buttons()


def cisco_onl_diff_compare():
    connect_db()
    c.execute('''
    SELECT DISTINCT mac_before, state_before, state_after
    FROM cisco_scm_core_before_full
    LEFT JOIN cisco_scm_core_after_full
    ON cisco_scm_core_before_full.mac_before = cisco_scm_core_after_full.mac_after
    WHERE (state_before = 'w-online(pt)' OR state_before = 'online(pt)') 
    AND (state_after != 'w-online(pt)' AND state_after != 'online(pt)' OR state_after IS NULL)
    ORDER BY state_after,state_before
    ''')
    rows_compare = c.fetchall()
    output = ""
    for row in rows_compare:
        mac_before = row[0] if row[0] is not None else "None"
        state_before = row[1] if row[1] is not None else "None"
        state_after = row[2] if row[2] is not None else "None"
        
        output += f"{mac_before:<16} {state_before:<12} {state_after:<12}\n"


    query_output_box.delete(1.0, END)
    query_output_box.insert(END, output)
    cisco_status_label.configure(text = 'CISCO_ONLINE_BEFORE_LOADED',fg_color='#054FA8')
    casa_status_label.configure(text = '', fg_color='transparent')
    disable_xls_buttons()

def cisco_state_diff_compare():
    connect_db()
    c.execute('''
    SELECT DISTINCT mac_before, state_before, state_after
    FROM cisco_scm_core_before_full
    LEFT JOIN cisco_scm_core_after_full
    ON cisco_scm_core_before_full.mac_before = cisco_scm_core_after_full.mac_after
    WHERE (state_before != state_after) OR state_after IS NULL
    ORDER BY state_after,state_before
    ''')
    rows_compare = c.fetchall()
    output = ""
    for row in rows_compare:
        mac_before = row[0] if row[0] is not None else "None"
        state_before = row[1] if row[1] is not None else "None"
        state_after = row[2] if row[2] is not None else "None"
        
        output += f"{mac_before:<16} {state_before:<15} {state_after:<12}\n"


    query_output_box.delete(1.0, END)
    query_output_box.insert(END, output)
    cisco_compare_new_modems_correction()
    cisco_status_label.configure(text = 'CISCO_BEFORE!=AFTER_LOADED',fg_color='#054FA8')
    casa_status_label.configure(text = '', fg_color='transparent')
    disable_xls_buttons()
    
def cisco_scm_detail():
    connect_db()
    c.execute('''
        SELECT DISTINCT mac_before, state_before, state_after,interface_before,interface_after
        FROM cisco_scm_core_before_full
        LEFT JOIN cisco_scm_core_after_full
        ON cisco_scm_core_before_full.mac_before = cisco_scm_core_after_full.mac_after
        ORDER BY state_after,state_before
        ''')
    
    rows_compare = c.fetchall()
    output = ""
    for row in rows_compare:
        mac_before = row[0] if row[0] is not None else "None"
        state_before = row[1] if row[1] is not None else "None"
        state_after = row[2] if row[2] is not None else "None"
        interface_before = row[3] if row[3] is not None else "None"
        interface_after = row[4] if row[4] is not None else "None"
        output += f"{mac_before:<16} {state_before:<15} {state_after:<15}{interface_before:<15}{interface_after:<15}\n"

    scm_before_input.delete(1.0, END)
    scm_before_input.insert(END, output)
    cisco_status_label.configure(text = 'CISCO_DETAIL_LOADED',fg_color='#054FA8')
    casa_status_label.configure(text = '', fg_color='transparent')
    enable_cisco_xls_button()
    disable_casa_xls_button()

def export_xl_cisco():

    current_folder = os.getcwd() 
    file_path = os.path.join(current_folder, "cisco_analysis.xlsx") 
    
    rows = scm_before_input.get(1.0, "end-1c").splitlines()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'CISCO CMTS MODEM ANALYSIS'

    column_names = ["MAC", "STATUS_BEFORE", "STATUS_AFTER",'INTERFACE_BEFORE','INTERFACE_AFTER']
    for col_index, col_name in enumerate(column_names, start=1):
        col_letter = get_column_letter(col_index)
        sheet[col_letter + '1'] = col_name
        sheet[col_letter + '1'].font = Font(bold=True)

    for row_index, row in enumerate(rows, start=2):
        columns = row.split()
        for col_index, col_value in enumerate(columns, start=1):
            col_letter = get_column_letter(col_index)
            sheet[col_letter + str(row_index)] = col_value
            
    sheet.column_dimensions['A'].width = 18  
    sheet.column_dimensions['B'].width = 20          
    sheet.column_dimensions['C'].width = 20  
    
    sheet.column_dimensions['D'].width = 22   
    sheet.column_dimensions['E'].width = 22   
    
    sheet.column_dimensions['F'].width = 20  
    sheet.column_dimensions['I'].width = 20  
    
    sheet.column_dimensions['G'].width = 14  
    sheet.column_dimensions['J'].width = 14  
    
    sheet['F2'] = 'TOTAL_BEFORE'
    sheet['F3'] = 'W-ONLINE_BEFORE'
    sheet['F4'] = 'P-ONLINE_BEFORE'
    sheet['F5'] = 'ONLINE(PT)_BEFORE'
    sheet['F6'] = 'OFFLINE_BEFORE'
    sheet['F7'] = 'INIT_BEFORE'
    
    sheet['I2'] = 'TOTAL_AFTER'
    sheet['I3'] = 'W-ONLINE_AFTER'
    sheet['I4'] = 'P-ONLINE_AFTER'
    sheet['I5'] = 'ONLINE(PT)_AFTER'
    sheet['I6'] = 'OFFLINE_AFTER'
    sheet['I7'] = 'INIT_AFTER'
    
    count_total_formula_pre = '=COUNTIF(B2:B{}, "<>None")'.format(len(rows) + 1)
    sheet['G2'] = count_total_formula_pre
    
    count_wonline_formula_pre = '=COUNTIF(B2:B{}, "w-online(pt)")'.format(len(rows) + 1)
    sheet['G3'] = count_wonline_formula_pre
    
    count_ponline_formula_pre = '=COUNTIF(B2:B{}, "p-online(pt)")'.format(len(rows) + 1)
    sheet['G4'] = count_ponline_formula_pre
    
    count_online_formula_pre = '=COUNTIF(B2:B{}, "online(pt)")'.format(len(rows) + 1)
    sheet['G5'] = count_online_formula_pre
    
    count_offline_formula_pre = '=COUNTIF(B2:B{}, "offline")'.format(len(rows) + 1)
    sheet['G6'] = count_offline_formula_pre
 
    count_init_formula_pre = '=COUNTIF(B2:B{}, "init*")'.format(len(rows) + 1)
    sheet['G7'] = count_init_formula_pre
    
    count_total_formula_posle = '=COUNTIF(C2:C{}, "<>None")'.format(len(rows) + 1)
    sheet['J2'] = count_total_formula_posle
    
    count_wonline_formula_posle = '=COUNTIF(C2:C{}, "w-online(pt)")'.format(len(rows) + 1)
    sheet['J3'] = count_wonline_formula_posle
    
    count_ponline_formula_posle = '=COUNTIF(C2:C{}, "p-online(pt)")'.format(len(rows) + 1)
    sheet['J4'] = count_ponline_formula_posle
    
    count_online_formula_posle = '=COUNTIF(C2:C{}, "online(pt)")'.format(len(rows) + 1)
    sheet['J5'] = count_online_formula_posle
    
    count_offline_formula_posle = '=COUNTIF(C2:C{}, "offline")'.format(len(rows) + 1)
    sheet['J6'] = count_offline_formula_posle
    
    count_init_formula_posle = '=COUNTIF(C2:C{}, "init*")'.format(len(rows) + 1)
    sheet['J7'] = count_init_formula_posle

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    fill = PatternFill(fill_type='solid', fgColor='E5F993')
    for row in sheet.iter_rows(min_row=2, max_row=7, min_col=6, max_col=7):
        for cell in row:
            cell.fill = fill
            cell.border = border
            
    fill = PatternFill(fill_type='solid', fgColor='F9DC5C')
    for row in sheet.iter_rows(min_row=2, max_row=7, min_col=9, max_col=10):
        for cell in row:
            cell.fill = fill
            cell.border = border
    
    try:
        os.remove(file_path)
        workbook.save(file_path)
        messagebox.showinfo('EXPORT SUCCESS', "OUTPUT SAVED TO cisco_analysis.xlsx")
        
    except Exception as e:
        workbook.save(file_path)
        messagebox.showinfo('EXPORT SUCCESS', "OUTPUT SAVED TO cisco_analysis.xlsx")
        pass
    
    os.startfile(file_path)
    
    pass

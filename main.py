from cisco_cbr_class import CiscoScm
from casa_cmts_class import CasaScm
import atexit
import sqlite3
import os
import xlwt
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side
import logging
#######################################################################


logging.basicConfig(level=logging.INFO, filename = 'log.log',filemode="w",
                    format= "%(asctime)s - %(levelname)s - %(message)s")


#######################################################################
from interface_stage import *
root.mainloop()